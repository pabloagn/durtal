"""Step 6: Seed series from Book_Series sheet and reconcile works.series_name → series_id."""

import openpyxl
from rich.console import Console

from scripts.ingest.config import EXCEL_PATH
from scripts.ingest.db import transaction, upsert_returning_id, lookup_id_ilike
from scripts.ingest.utils import clean, slugify

console = Console()


def seed_series_from_sheet(cur, wb, *, dry_run: bool = False) -> int:
    """Seed series table from Book_Series sheet."""
    if "Book_Series" not in wb.sheetnames:
        console.print("[yellow]Sheet 'Book_Series' not found — skipping[/yellow]")
        return 0

    ws = wb["Book_Series"]
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Title, Original Title, Author, Composite_Work_Author, Composite_Author_Work, Duplicated_Entry
        title = clean(row[0])
        original_title = clean(row[1]) if len(row) > 1 else None
        is_dup = clean(row[5]) if len(row) > 5 else None

        if not title or is_dup == "Y":
            continue

        slug = slugify(title)
        if not slug:
            continue

        if dry_run:
            count += 1
            continue

        upsert_returning_id(
            cur, "series",
            ["title", "original_title", "slug"],
            (title, original_title, slug),
            "slug",
            update_on_conflict=True,
            update_columns=["original_title"],
        )
        count += 1

    return count


def reconcile_series(cur, *, dry_run: bool = False) -> dict:
    """Find works with series_name set but series_id NULL, and link them."""
    stats = {"reconciled": 0, "series_created": 0}

    cur.execute(
        "SELECT id, series_name FROM works WHERE series_name IS NOT NULL AND series_id IS NULL"
    )
    orphans = cur.fetchall()

    if not orphans:
        return stats

    console.print(f"  Found {len(orphans)} works with series_name but no series_id")

    if dry_run:
        stats["reconciled"] = len(orphans)
        return stats

    for work_id, series_name in orphans:
        name = clean(series_name)
        if not name:
            continue

        series_id = lookup_id_ilike(cur, "series", "title", name)

        if not series_id:
            slug = slugify(name)
            series_id = upsert_returning_id(
                cur, "series",
                ["title", "slug"],
                (name, slug),
                "slug",
            )
            stats["series_created"] += 1

        if series_id:
            cur.execute(
                "UPDATE works SET series_id = %s WHERE id = %s",
                (series_id, work_id),
            )
            stats["reconciled"] += 1

    return stats


def run(*, dry_run: bool = False):
    console.rule("[bold]Step 6: Series")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    try:
        if dry_run:
            console.print("[yellow]DRY RUN — no data will be written[/yellow]\n")
            seeded = seed_series_from_sheet(None, wb, dry_run=True)
            console.print(f"  series from sheet: {seeded}")
            console.print("  reconciliation:    (skipped in dry run)")
        else:
            with transaction() as cur:
                seeded = seed_series_from_sheet(cur, wb)
                stats = reconcile_series(cur)
            console.print(f"  series from sheet: {seeded}")
            console.print(f"  works reconciled:  {stats['reconciled']}")
            console.print(f"  series created:    {stats['series_created']}")

        console.print("[green]Series complete.[/green]\n")
    finally:
        wb.close()


if __name__ == "__main__":
    import sys
    run(dry_run="--dry-run" in sys.argv)
