"""Step 2: Seed taxonomy tables — hierarchical (book_categories, literary_movements, themes) and flat (art_types, art_movements, keywords, attributes, subjects)."""

import openpyxl
from rich.console import Console

from scripts.ingest.config import EXCEL_PATH
from scripts.ingest.db import transaction, upsert_returning_id
from scripts.ingest.utils import clean, slugify, slugify_path

console = Console()


def _load_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    console.print(f"[yellow]Sheet '{name}' not found — skipping[/yellow]")
    return None


# ── Hierarchical seeding ─────────────────────────────────────────────────────


def _seed_hierarchy(
    cur,
    wb,
    sheet_name: str,
    table_name: str,
    col_indices: tuple[int, int, int],
    *,
    dup_col: int | None = None,
    scope_col: int | None = None,
    dry_run: bool = False,
) -> int:
    """Generic 3-level hierarchy seeder.

    col_indices: tuple of (level1_col, level2_col, level3_col) indices.
    dup_col: optional column index for Duplicated_Entry flag.
    scope_col: optional column index for scope_notes / notes.
    """
    ws = _load_sheet(wb, sheet_name)
    if not ws:
        return 0

    # Build the tree: {l1: {l2: {l3: scope_notes}}}
    tree: dict[str, dict[str, dict[str, str | None]]] = {}
    scope_notes_map: dict[str, str | None] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if dup_col is not None and len(row) > dup_col and clean(row[dup_col]) == "Y":
            continue

        l1 = clean(row[col_indices[0]]) if len(row) > col_indices[0] else None
        l2 = clean(row[col_indices[1]]) if len(row) > col_indices[1] else None
        l3 = clean(row[col_indices[2]]) if len(row) > col_indices[2] else None
        scope = clean(row[scope_col]) if scope_col is not None and len(row) > scope_col else None

        if not l1:
            continue

        if l1 not in tree:
            tree[l1] = {}
        if scope and l3:
            scope_notes_map[f"{l1}/{l2}/{l3}"] = scope
        elif scope and l2:
            scope_notes_map[f"{l1}/{l2}"] = scope
        elif scope:
            scope_notes_map[l1] = scope

        if l2:
            if l2 not in tree[l1]:
                tree[l1][l2] = {}
            if l3:
                tree[l1][l2][l3] = scope

    if dry_run:
        # Count total nodes
        count = 0
        for l1, children in tree.items():
            count += 1  # l1
            for l2, grandchildren in children.items():
                count += 1  # l2
                count += len(grandchildren)  # l3s
        return count

    count = 0
    sort_order = 0

    # Has scope_notes column?
    has_scope = table_name in ("book_categories", "literary_movements")
    scope_cols = ", scope_notes" if has_scope else ""
    scope_placeholder = ", %s" if has_scope else ""

    # Insert level 1 (roots)
    l1_ids: dict[str, str] = {}
    for l1_name in tree:
        sort_order += 1
        slug = slugify(l1_name)
        scope = scope_notes_map.get(l1_name)

        if has_scope:
            cur.execute(
                f"""INSERT INTO {table_name} (name, slug, level, parent_id, sort_order{scope_cols})
                VALUES (%s, %s, 1, NULL, %s{scope_placeholder})
                ON CONFLICT (slug) DO NOTHING RETURNING id""",
                (l1_name, slug, sort_order, scope) if has_scope else (l1_name, slug, sort_order),
            )
        else:
            cur.execute(
                f"""INSERT INTO {table_name} (name, slug, level, parent_id, sort_order)
                VALUES (%s, %s, 1, NULL, %s)
                ON CONFLICT (slug) DO NOTHING RETURNING id""",
                (l1_name, slug, sort_order),
            )

        row = cur.fetchone()
        if row:
            l1_ids[l1_name] = row[0]
        else:
            cur.execute(f"SELECT id FROM {table_name} WHERE slug = %s", (slug,))
            r = cur.fetchone()
            l1_ids[l1_name] = r[0] if r else None
        count += 1

    # Insert level 2
    l2_ids: dict[str, str] = {}
    for l1_name, children in tree.items():
        parent_id = l1_ids.get(l1_name)
        for l2_name in children:
            sort_order += 1
            slug = slugify_path([l1_name, l2_name])
            scope = scope_notes_map.get(f"{l1_name}/{l2_name}")

            if has_scope:
                cur.execute(
                    f"""INSERT INTO {table_name} (name, slug, level, parent_id, sort_order{scope_cols})
                    VALUES (%s, %s, 2, %s, %s{scope_placeholder})
                    ON CONFLICT (slug) DO NOTHING RETURNING id""",
                    (l2_name, slug, parent_id, sort_order, scope) if has_scope else (l2_name, slug, parent_id, sort_order),
                )
            else:
                cur.execute(
                    f"""INSERT INTO {table_name} (name, slug, level, parent_id, sort_order)
                    VALUES (%s, %s, 2, %s, %s)
                    ON CONFLICT (slug) DO NOTHING RETURNING id""",
                    (l2_name, slug, parent_id, sort_order),
                )

            row = cur.fetchone()
            key = f"{l1_name}/{l2_name}"
            if row:
                l2_ids[key] = row[0]
            else:
                cur.execute(f"SELECT id FROM {table_name} WHERE slug = %s", (slug,))
                r = cur.fetchone()
                l2_ids[key] = r[0] if r else None
            count += 1

    # Insert level 3
    for l1_name, children in tree.items():
        for l2_name, grandchildren in children.items():
            parent_key = f"{l1_name}/{l2_name}"
            parent_id = l2_ids.get(parent_key)
            for l3_name in grandchildren:
                sort_order += 1
                slug = slugify_path([l1_name, l2_name, l3_name])
                scope = scope_notes_map.get(f"{l1_name}/{l2_name}/{l3_name}")

                if has_scope:
                    cur.execute(
                        f"""INSERT INTO {table_name} (name, slug, level, parent_id, sort_order{scope_cols})
                        VALUES (%s, %s, 3, %s, %s{scope_placeholder})
                        ON CONFLICT (slug) DO NOTHING RETURNING id""",
                        (l3_name, slug, parent_id, sort_order, scope) if has_scope else (l3_name, slug, parent_id, sort_order),
                    )
                else:
                    cur.execute(
                        f"""INSERT INTO {table_name} (name, slug, level, parent_id, sort_order)
                        VALUES (%s, %s, 3, %s, %s)
                        ON CONFLICT (slug) DO NOTHING RETURNING id""",
                        (l3_name, slug, parent_id, sort_order),
                    )
                count += 1

    return count


def seed_book_categories(cur, wb, *, dry_run: bool = False) -> int:
    return _seed_hierarchy(
        cur, wb, "Book_Categories", "book_categories",
        col_indices=(0, 1, 2),  # Category_1, Category_2, Category_3
        scope_col=3,            # Notes_Scope_Example
        dry_run=dry_run,
    )


def seed_literary_movements(cur, wb, *, dry_run: bool = False) -> int:
    return _seed_hierarchy(
        cur, wb, "Literary_Movements", "literary_movements",
        col_indices=(0, 1, 2),  # Level 1, Level 2, Level 3
        dup_col=4,              # Duplicated_Entry
        scope_col=3,            # Notes / Scope Example
        dry_run=dry_run,
    )


def seed_themes(cur, wb, *, dry_run: bool = False) -> int:
    return _seed_hierarchy(
        cur, wb, "Themes", "themes",
        col_indices=(0, 1, 2),  # Category, Subcategory, Theme
        dry_run=dry_run,
    )


# ── Flat table seeding ───────────────────────────────────────────────────────


def seed_art_types(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Art_Types")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Book_Type, Description, Applicable_Work_Type(s), Duplicated_Entry
        name = clean(row[0])
        description = clean(row[1]) if len(row) > 1 else None
        applicable = clean(row[2]) if len(row) > 2 else None
        is_dup = clean(row[3]) if len(row) > 3 else None

        if not name or is_dup == "Y":
            continue

        slug = slugify(name)
        if dry_run:
            count += 1
            continue
        upsert_returning_id(
            cur, "art_types",
            ["name", "slug", "description", "applicable_work_types"],
            (name, slug, description, applicable),
            "name",
        )
        count += 1
    return count


def seed_art_movements(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Art_Movements")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[0])
        is_dup = clean(row[1]) if len(row) > 1 else None
        if not name or is_dup == "Y":
            continue
        slug = slugify(name)
        if dry_run:
            count += 1
            continue
        upsert_returning_id(cur, "art_movements", ["name", "slug"], (name, slug), "name")
        count += 1
    return count


def seed_keywords(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Keywords")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[0])
        if not name:
            continue
        slug = slugify(name)
        if dry_run:
            count += 1
            continue
        upsert_returning_id(cur, "keywords", ["name", "slug"], (name, slug), "name")
        count += 1
    return count


def seed_attributes(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Attributes")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[0])
        description = clean(row[1]) if len(row) > 1 else None
        category = clean(row[2]) if len(row) > 2 else None
        is_dup = clean(row[3]) if len(row) > 3 else None
        if not name or is_dup == "Y":
            continue
        slug = slugify(name)
        if dry_run:
            count += 1
            continue
        upsert_returning_id(
            cur, "attributes",
            ["name", "slug", "description", "category"],
            (name, slug, description, category),
            "name",
        )
        count += 1
    return count


def seed_subjects(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Subjects")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Name, Description (Optional)
        name = clean(row[0])
        description = clean(row[1]) if len(row) > 1 else None
        if not name:
            continue
        slug = slugify(name)
        if dry_run:
            count += 1
            continue
        upsert_returning_id(
            cur, "subjects",
            ["name", "slug", "description"],
            (name, slug, description),
            "name",
            update_on_conflict=True,
            update_columns=["description"],
        )
        count += 1
    return count


def run(*, dry_run: bool = False):
    """Execute all taxonomy seeding."""
    console.rule("[bold]Step 2: Taxonomy Data")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    try:
        if dry_run:
            console.print("[yellow]DRY RUN — no data will be written[/yellow]\n")
            bc = seed_book_categories(None, wb, dry_run=True)
            lm = seed_literary_movements(None, wb, dry_run=True)
            th = seed_themes(None, wb, dry_run=True)
            at = seed_art_types(None, wb, dry_run=True)
            am = seed_art_movements(None, wb, dry_run=True)
            kw = seed_keywords(None, wb, dry_run=True)
            ab = seed_attributes(None, wb, dry_run=True)
            su = seed_subjects(None, wb, dry_run=True)
        else:
            with transaction() as cur:
                bc = seed_book_categories(cur, wb)
                lm = seed_literary_movements(cur, wb)
                th = seed_themes(cur, wb)
                at = seed_art_types(cur, wb)
                am = seed_art_movements(cur, wb)
                kw = seed_keywords(cur, wb)
                ab = seed_attributes(cur, wb)
                su = seed_subjects(cur, wb)

        console.print(f"  book_categories:    {bc}")
        console.print(f"  literary_movements: {lm}")
        console.print(f"  themes:             {th}")
        console.print(f"  art_types:          {at}")
        console.print(f"  art_movements:      {am}")
        console.print(f"  keywords:           {kw}")
        console.print(f"  attributes:         {ab}")
        console.print(f"  subjects:           {su}")
        console.print("[green]Taxonomy data complete.[/green]\n")

    finally:
        wb.close()


if __name__ == "__main__":
    import sys
    run(dry_run="--dry-run" in sys.argv)
