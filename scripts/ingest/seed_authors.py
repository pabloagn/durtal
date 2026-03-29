"""Step 4: Seed authors from People sheet, filtered to literary roles."""

import openpyxl
from rich.console import Console

from scripts.ingest.config import EXCEL_PATH, NON_LITERARY_ROLES, METADATA_SOURCE, NATIONALITY_TO_COUNTRY, ROLE_NORMALIZATION
from scripts.ingest.db import (
    transaction,
    upsert_returning_id,
    bulk_load_lookup,
    insert_junction,
)
from scripts.ingest.utils import clean, clean_int, slugify, parse_roles

console = Console()

# Column indices for People sheet
# Hash_ID(0), Name(1), Surname(2), Real Name(3), Type(4), Gender(5), Nationality(6),
# Birth_Date_Day(7), Birth_Date_Month(8), Birth_Date_Year(9), Birth_Date_Year_Gregorian(10),
# Birth_Date_Year_IsRange(11), Death_Date_Day(12), Death_Date_Month(13), Death_Date_Year(14),
# Death_Date_Year_Gregorian(15), Death_Date_Year_IsRange(16),
# Complete Name (Name Surname)(17), Complete Name (Surname Name)(18),
# Duplicated_Entry(19), Has_Image(20)


def _is_literary(roles: list[str]) -> bool:
    """Return True if the person has at least one role NOT in the non-literary set."""
    if not roles:
        return False
    return any(r not in NON_LITERARY_ROLES for r in roles)


def _build_nationality_resolver(cur) -> dict[str, str]:
    """Build a nationality → country_id mapping using config + DB countries."""
    # Load all countries from DB
    cur.execute("SELECT id, name FROM countries")
    countries_rows = cur.fetchall()

    # Index by lowercase name and prefix
    name_to_id: dict[str, str] = {}
    for row in countries_rows:
        name_to_id[row[1].lower()] = row[0]

    # Build nationality → country_id
    nat_to_id: dict[str, str] = {}
    for nat, country_search in NATIONALITY_TO_COUNTRY.items():
        search_lower = country_search.lower()
        # Try exact match first
        if search_lower in name_to_id:
            nat_to_id[nat] = name_to_id[search_lower]
            continue
        # Try prefix match
        for cn, cid in name_to_id.items():
            if cn.startswith(search_lower):
                nat_to_id[nat] = cid
                break

    return nat_to_id


def seed_authors(cur, wb, *, dry_run: bool = False, limit: int | None = None) -> dict:
    ws = wb["People"] if "People" in wb.sheetnames else None
    if not ws:
        console.print("[yellow]Sheet 'People' not found — skipping[/yellow]")
        return {"ingested": 0, "skipped_dup": 0, "skipped_non_literary": 0}

    # Pre-load contribution_types lookup
    ct_lookup = bulk_load_lookup(cur, "contribution_types") if not dry_run else {}

    # Build nationality resolver
    nat_to_id: dict[str, str] = {}
    if not dry_run:
        nat_to_id = _build_nationality_resolver(cur)

    stats = {"ingested": 0, "skipped_dup": 0, "skipped_non_literary": 0, "unmatched_types": set()}
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        first_name = clean(row[1])
        last_name = clean(row[2])
        real_name = clean(row[3])
        type_raw = clean(row[4])
        gender = clean(row[5])
        nationality = clean(row[6])
        birth_day = clean_int(row[7])
        birth_month = clean_int(row[8])
        birth_year = clean_int(row[9])
        birth_year_gregorian = clean_int(row[10])
        birth_year_is_approx = clean(row[11])
        death_day = clean_int(row[12])
        death_month = clean_int(row[13])
        death_year = clean_int(row[14])
        death_year_gregorian = clean_int(row[15])
        death_year_is_approx = clean(row[16])
        is_dup = clean(row[19]) if len(row) > 19 else None

        # Skip duplicates
        if is_dup == "Y":
            stats["skipped_dup"] += 1
            continue

        # Parse roles (split on both commas and semicolons)
        roles = parse_roles(type_raw) if type_raw else []

        # Filter: must have at least one literary role
        if not _is_literary(roles):
            stats["skipped_non_literary"] += 1
            continue

        # Build display name and sort name
        if first_name and last_name:
            display_name = f"{first_name} {last_name}"
            sort_name = f"{last_name}, {first_name}"
        elif first_name:
            display_name = first_name
            sort_name = first_name
        elif last_name:
            display_name = last_name
            sort_name = last_name
        else:
            continue  # no usable name

        slug = slugify(display_name)
        if not slug:
            continue

        # Gender mapping
        gender_val = None
        if gender:
            g = gender.lower().strip()
            if g in ("male", "m"):
                gender_val = "male"
            elif g in ("female", "f"):
                gender_val = "female"

        # Nationality → country_id (handle compound like "Austrian; American" — take first)
        nationality_id = None
        if nationality and not dry_run:
            first_nat = nationality.split(";")[0].strip()
            nationality_id = nat_to_id.get(first_nat)

        # Approximate flags
        birth_approx = birth_year_is_approx in ("Y", "Yes", "TRUE", "True", "1") if birth_year_is_approx else False
        death_approx = death_year_is_approx in ("Y", "Yes", "TRUE", "True", "1") if death_year_is_approx else False

        if dry_run:
            stats["ingested"] += 1
            count += 1
            if limit and count >= limit:
                break
            continue

        author_id = upsert_returning_id(
            cur, "authors",
            [
                "name", "slug", "sort_name", "first_name", "last_name", "real_name",
                "gender", "birth_year", "birth_month", "birth_day",
                "birth_year_is_approximate", "birth_year_gregorian",
                "death_year", "death_month", "death_day",
                "death_year_is_approximate", "death_year_gregorian",
                "nationality_id", "metadata_source",
            ],
            (
                display_name, slug, sort_name, first_name, last_name, real_name,
                gender_val, birth_year, birth_month, birth_day,
                birth_approx, birth_year_gregorian,
                death_year, death_month, death_day,
                death_approx, death_year_gregorian,
                nationality_id, METADATA_SOURCE,
            ),
            "slug",
            update_on_conflict=True,
            update_columns=[
                "sort_name", "first_name", "last_name", "real_name",
                "gender", "birth_year", "birth_month", "birth_day",
                "birth_year_is_approximate", "birth_year_gregorian",
                "death_year", "death_month", "death_day",
                "death_year_is_approximate", "death_year_gregorian",
                "nationality_id",
            ],
        )

        # Tag with contribution types
        if author_id:
            ct_lower = {k.lower(): v for k, v in ct_lookup.items()}
            for role in roles:
                ct_id = ct_lookup.get(role)
                if not ct_id:
                    ct_id = ct_lower.get(role.lower())
                if not ct_id:
                    # Try role normalization map
                    normalized = ROLE_NORMALIZATION.get(role)
                    if normalized:
                        ct_id = ct_lookup.get(normalized) or ct_lower.get(normalized.lower())
                if ct_id:
                    insert_junction(cur, "author_contribution_types", "author_id", "contribution_type_id", author_id, ct_id)
                else:
                    stats["unmatched_types"].add(role)

        stats["ingested"] += 1
        count += 1
        if limit and count >= limit:
            break

    return stats


def run(*, dry_run: bool = False, limit: int | None = None):
    console.rule("[bold]Step 4: Authors")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    try:
        if dry_run:
            console.print("[yellow]DRY RUN — no data will be written[/yellow]\n")
            stats = seed_authors(None, wb, dry_run=True, limit=limit)
        else:
            with transaction() as cur:
                stats = seed_authors(cur, wb, limit=limit)

        console.print(f"  authors ingested:      {stats['ingested']}")
        console.print(f"  skipped (duplicate):   {stats['skipped_dup']}")
        console.print(f"  skipped (non-literary): {stats['skipped_non_literary']}")
        if stats.get("unmatched_types"):
            console.print(f"  [yellow]unmatched contribution types: {sorted(stats['unmatched_types'])}[/yellow]")
        console.print("[green]Authors complete.[/green]\n")

    finally:
        wb.close()


if __name__ == "__main__":
    import sys
    run(
        dry_run="--dry-run" in sys.argv,
        limit=int(sys.argv[sys.argv.index("--limit") + 1]) if "--limit" in sys.argv else None,
    )
