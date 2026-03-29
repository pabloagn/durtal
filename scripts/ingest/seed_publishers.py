"""Step 3: Seed publishing houses, publisher specialties, and their junction."""

import openpyxl
from rich.console import Console

from scripts.ingest.config import EXCEL_PATH
from scripts.ingest.db import (
    transaction,
    upsert_returning_id,
    bulk_load_lookup,
    lookup_id_ilike,
    insert_junction,
)
from scripts.ingest.utils import clean, slugify, parse_semicolons

console = Console()


def _load_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    console.print(f"[yellow]Sheet '{name}' not found — skipping[/yellow]")
    return None


def seed_publisher_specialties(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Publisher_Specialties")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[0])
        if not name:
            continue
        slug = slugify(name)
        if dry_run:
            count += 1
            continue
        upsert_returning_id(cur, "publisher_specialties", ["name", "slug"], (name, slug), "name")
        count += 1
    return count


def seed_publishing_houses(cur, wb, *, dry_run: bool = False) -> tuple[int, int]:
    """Seed publishing_houses and junction. Returns (houses_count, junction_count)."""
    ws = _load_sheet(wb, "Publishing_Houses")
    if not ws:
        return 0, 0

    # Pre-load lookups
    specialty_lookup = bulk_load_lookup(cur, "publisher_specialties") if not dry_run else {}

    houses_count = 0
    junction_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Publishing_House, Country, Duplicated_Entry, Specialties
        name = clean(row[0])
        country_text = clean(row[1])
        is_dup = clean(row[2]) if len(row) > 2 else None
        specialties_raw = clean(row[3]) if len(row) > 3 else None

        if not name or is_dup == "Y":
            continue

        slug = slugify(name)

        if dry_run:
            houses_count += 1
            continue

        # Resolve country → country_id
        country_id = None
        if country_text:
            country_id = lookup_id_ilike(cur, "countries", "name", country_text)
            if not country_id:
                # Try partial match: "United States" → "United States of America"
                cur.execute(
                    "SELECT id FROM countries WHERE LOWER(name) LIKE %s LIMIT 1",
                    (f"%{country_text.lower()}%",),
                )
                r = cur.fetchone()
                country_id = r[0] if r else None

        house_id = upsert_returning_id(
            cur, "publishing_houses",
            ["name", "slug", "country", "country_id"],
            (name, slug, country_text, country_id),
            "name",
        )
        houses_count += 1

        # Link specialties
        if specialties_raw and house_id:
            parts = parse_semicolons(specialties_raw)
            for sp_name in parts:
                sp_id = specialty_lookup.get(sp_name)
                if sp_id:
                    insert_junction(cur, "publishing_house_specialties", "publishing_house_id", "specialty_id", house_id, sp_id)
                    junction_count += 1

    return houses_count, junction_count


def run(*, dry_run: bool = False):
    console.rule("[bold]Step 3: Publishing Houses")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    try:
        if dry_run:
            console.print("[yellow]DRY RUN — no data will be written[/yellow]\n")
            sp = seed_publisher_specialties(None, wb, dry_run=True)
            console.print(f"  publisher_specialties: {sp}")
            console.print("  publishing_houses:    (skipped in dry run)")
            console.print("[green]Publishing houses dry run complete.[/green]\n")
        else:
            with transaction() as cur:
                sp = seed_publisher_specialties(cur, wb)
                houses, junctions = seed_publishing_houses(cur, wb)
            console.print(f"  publisher_specialties:          {sp}")
            console.print(f"  publishing_houses:              {houses}")
            console.print(f"  publishing_house_specialties:   {junctions}")
            console.print("[green]Publishing houses complete.[/green]\n")

    finally:
        wb.close()


if __name__ == "__main__":
    import sys
    run(dry_run="--dry-run" in sys.argv)
