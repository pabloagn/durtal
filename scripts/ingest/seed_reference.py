"""Step 1: Seed reference/lookup tables — languages, countries, centuries, work_types, contribution_types, sources."""

import openpyxl
from rich.console import Console

from scripts.ingest.config import EXCEL_PATH
from scripts.ingest.db import transaction, upsert_returning_id
from scripts.ingest.utils import clean, clean_int, slugify, parse_century

console = Console()


def _load_sheet(wb, name: str):
    """Return a worksheet by name, or None if missing."""
    if name in wb.sheetnames:
        return wb[name]
    console.print(f"[yellow]Sheet '{name}' not found — skipping[/yellow]")
    return None


def seed_languages(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Languages")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Hash_ID, Language, ISO_639-1, ISO_639-1_Is_Duplicate, ISO_639-2, ISO_639-2_Is_Duplicate, ISO_639-3, ISO_639-3_Is_Duplicate
        name = clean(row[1])
        iso1 = clean(row[2])
        is_dup = clean(row[3])
        iso2 = clean(row[4])
        iso3 = clean(row[6])

        if not name or is_dup == "Y":
            continue

        if dry_run:
            count += 1
            continue

        upsert_returning_id(
            cur, "languages",
            ["name", "iso_639_1", "iso_639_2", "iso_639_3"],
            (name, iso1, iso2, iso3),
            "name",
        )
        count += 1

    return count


def seed_countries(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Countries_Continents")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Continent_Name, Continent_Code, Country_Name, Two_Letter_Country_Code, Three_Letter_Country_Code, Country_Number, ...
        continent_name = clean(row[0])
        continent_code = clean(row[1])
        country_name = clean(row[2])
        alpha2 = clean(row[3])
        alpha3 = clean(row[4])
        numeric_code = clean_int(row[5])

        if not country_name or not alpha2 or not alpha3:
            continue

        if dry_run:
            count += 1
            continue

        upsert_returning_id(
            cur, "countries",
            ["name", "alpha_2", "alpha_3", "numeric_code", "continent_name", "continent_code"],
            (country_name, alpha2, alpha3, numeric_code, continent_name, continent_code),
            "name",
        )
        count += 1

    return count


def seed_centuries(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Validation_Fields")
    if not ws:
        return 0

    count = 0
    sort_order = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Century labels are in column H (index 7)
        label = clean(row[7]) if len(row) > 7 else None
        if not label or "Century" not in label:
            continue

        start_year, end_year = parse_century(label)
        if start_year == 0:
            continue

        sort_order += 1

        if dry_run:
            count += 1
            continue

        upsert_returning_id(
            cur, "centuries",
            ["label", "start_year", "end_year", "sort_order"],
            (label, start_year, end_year, sort_order),
            "label",
        )
        count += 1

    return count


def seed_work_types(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Work_Types")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Name, Description
        name = clean(row[0])
        description = clean(row[1]) if len(row) > 1 else None

        if not name:
            continue

        slug = slugify(name)

        if dry_run:
            count += 1
            continue

        upsert_returning_id(
            cur, "work_types",
            ["name", "slug", "description"],
            (name, slug, description),
            "name",
        )
        count += 1

    return count


def seed_contribution_types(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "Contribution_Types")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Name, Description, Examples of Use
        name = clean(row[0])
        description = clean(row[1]) if len(row) > 1 else None
        examples = clean(row[2]) if len(row) > 2 else None

        if not name:
            continue

        slug = slugify(name)
        # Store "Examples of Use" in applicable_work_types
        applicable = examples

        if dry_run:
            count += 1
            continue

        upsert_returning_id(
            cur, "contribution_types",
            ["name", "slug", "description", "applicable_work_types"],
            (name, slug, description, applicable),
            "name",
        )
        count += 1

    return count


def seed_sources(cur, wb, *, dry_run: bool = False) -> int:
    ws = _load_sheet(wb, "URLs")
    if not ws:
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Concept, URL
        name = clean(row[0])
        url = clean(row[1]) if len(row) > 1 else None

        if not name:
            continue

        if dry_run:
            count += 1
            continue

        upsert_returning_id(
            cur, "sources",
            ["name", "url"],
            (name, url),
            "name",
        )
        count += 1

    return count


def run(*, dry_run: bool = False):
    """Execute all reference data seeding."""
    console.rule("[bold]Step 1: Reference Data")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    try:
        if dry_run:
            console.print("[yellow]DRY RUN — no data will be written[/yellow]\n")
            langs = seed_languages(None, wb, dry_run=True)
            countries = seed_countries(None, wb, dry_run=True)
            cents = seed_centuries(None, wb, dry_run=True)
            wtypes = seed_work_types(None, wb, dry_run=True)
            ctypes = seed_contribution_types(None, wb, dry_run=True)
            srcs = seed_sources(None, wb, dry_run=True)
        else:
            with transaction() as cur:
                langs = seed_languages(cur, wb)
                countries = seed_countries(cur, wb)
                cents = seed_centuries(cur, wb)
                wtypes = seed_work_types(cur, wb)
                ctypes = seed_contribution_types(cur, wb)
                srcs = seed_sources(cur, wb)

        console.print(f"  languages:          {langs}")
        console.print(f"  countries:          {countries}")
        console.print(f"  centuries:          {cents}")
        console.print(f"  work_types:         {wtypes}")
        console.print(f"  contribution_types: {ctypes}")
        console.print(f"  sources:            {srcs}")
        console.print("[green]Reference data complete.[/green]\n")

    finally:
        wb.close()


if __name__ == "__main__":
    import sys
    run(dry_run="--dry-run" in sys.argv)
