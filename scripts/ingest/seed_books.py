"""Step 5: Seed books from the Books sheet — works + editions + instances.

Filters to Priority 5, 4, 3 by default (configurable via --priority flag).
"""

import re
import openpyxl
from rich.console import Console

from scripts.ingest.config import EXCEL_PATH, LOCATION_MAP, METADATA_SOURCE
from scripts.ingest.db import (
    transaction,
    upsert_returning_id,
    lookup_id,
    lookup_id_ilike,
    bulk_load_lookup,
    insert_junction,
)
from scripts.ingest.utils import (
    clean,
    clean_int,
    slugify,
    invert_author_name,
    split_authors,
    extract_goodreads_id,
    parse_semicolons,
)

console = Console()

# Books sheet column indices (0-based)
COL = {
    "hash_id": 0,
    "duplicated_entry": 1,
    "title": 2,
    "author": 3,
    "series": 4,
    "series_number": 5,
    "published_date": 6,
    "edition": 7,
    "publisher": 8,
    "published_title": 9,
    "published_language": 10,
    "page_count": 11,
    "description_genai": 12,
    "recommended_by": 13,
    "priority": 14,
    "rating": 15,
    "read": 16,
    "started_date": 17,
    "finished_date": 18,
    "lib_physical_mex": 19,
    "lib_physical_eur": 20,
    "lib_digital_main": 21,
    "lib_mobile_kindle": 22,
    "lib_mobile_ipad": 23,
    "lib_mobile_iphone": 24,
    "goodreads_path": 25,
    "hardcover_id": 26,
    "hardcover_path": 27,
    "cover_raw": 28,
    "cover_ftd": 29,
    "published_date_decade": 30,
    "published_date_century": 31,
    "book_type": 32,
    "temporary_tags": 33,
    "tradition": 34,
    "notes": 35,
    "composite_work_author": 36,
    "composite_author_work": 37,
    "composite_work_author_clean": 38,
    "goodreads_url": 39,
    "hardcover_url": 40,
}

# Library column keys mapped to their Excel column indices
LIB_COLUMNS = {
    "Library_Physical_MEX": COL["lib_physical_mex"],
    "Library_Physical_EUR": COL["lib_physical_eur"],
    "Library_Digital_Main": COL["lib_digital_main"],
    "Library_Mobile_Kindle": COL["lib_mobile_kindle"],
    "Library_Mobile_iPad": COL["lib_mobile_ipad"],
    "Library_Mobile_iPhone": COL["lib_mobile_iphone"],
}


def _col(row, key: str):
    """Safely get a column value from a row tuple."""
    idx = COL[key]
    return row[idx] if idx < len(row) else None


def _parse_publication_year(val) -> int | None:
    """Extract a year from various date formats."""
    if val is None:
        return None
    s = str(val).strip()
    # Try direct integer
    try:
        y = int(float(s))
        if 100 < y < 2100:
            return y
    except (ValueError, TypeError):
        pass
    # Try extracting 4-digit year
    m = re.search(r"\b(1[0-9]{3}|20[0-9]{2})\b", s)
    return int(m.group(1)) if m else None


def _has_any_library(row) -> bool:
    """Check if any Library_* column is populated."""
    for lib_key, idx in LIB_COLUMNS.items():
        val = row[idx] if idx < len(row) else None
        if val is not None and str(val).strip() and str(val).strip().upper() != "N":
            return True
    return False


def _derive_catalogue_status(row, priority: int | None) -> str:
    """Derive catalogue_status from library columns and priority."""
    if _has_any_library(row):
        return "accessioned"
    if priority is not None:
        if priority >= 4:
            return "wanted"
        elif priority >= 3:
            return "shortlisted"
    return "tracked"


def _derive_acquisition_priority(priority: int | None) -> str:
    """Map numeric priority to acquisition_priority enum."""
    if priority is None:
        return "none"
    if priority >= 5:
        return "urgent"
    if priority >= 4:
        return "high"
    if priority >= 3:
        return "medium"
    if priority >= 2:
        return "low"
    return "none"


def _resolve_or_create_author(cur, author_sort_name: str, author_lookup: dict) -> str | None:
    """Find an existing author by sort_name or name, or create a minimal record."""
    if not author_sort_name:
        return None

    sort_key = author_sort_name.strip()
    display_name = invert_author_name(sort_key)

    # Try exact sort_name match
    author_id = author_lookup.get(sort_key)
    if author_id:
        return author_id

    # Try display name match
    author_id = author_lookup.get(display_name)
    if author_id:
        return author_id

    # Try DB lookup by sort_name
    author_id = lookup_id_ilike(cur, "authors", "sort_name", sort_key)
    if author_id:
        author_lookup[sort_key] = author_id
        return author_id

    # Try DB lookup by name
    author_id = lookup_id_ilike(cur, "authors", "name", display_name)
    if author_id:
        author_lookup[sort_key] = author_id
        return author_id

    # Create a minimal author record
    slug = slugify(display_name)
    if not slug:
        return None

    # Decompose name
    first_name = None
    last_name = None
    if "," in sort_key:
        parts = sort_key.split(",", 1)
        last_name = parts[0].strip()
        first_name = parts[1].strip()

    author_id = upsert_returning_id(
        cur, "authors",
        ["name", "slug", "sort_name", "first_name", "last_name", "metadata_source"],
        (display_name, slug, sort_key, first_name, last_name, METADATA_SOURCE),
        "slug",
    )
    if author_id:
        author_lookup[sort_key] = author_id
    return author_id


def _resolve_or_create_series(cur, series_title: str, series_lookup: dict) -> str | None:
    """Find or create a series by title."""
    if not series_title:
        return None

    title = series_title.strip()
    series_id = series_lookup.get(title.lower())
    if series_id:
        return series_id

    # DB lookup
    series_id = lookup_id_ilike(cur, "series", "title", title)
    if series_id:
        series_lookup[title.lower()] = series_id
        return series_id

    # Create
    slug = slugify(title)
    if not slug:
        return None

    series_id = upsert_returning_id(
        cur, "series",
        ["title", "slug"],
        (title, slug),
        "slug",
    )
    if series_id:
        series_lookup[title.lower()] = series_id
    return series_id


def _ensure_locations(cur) -> dict[str, str]:
    """Ensure all required locations exist and return {location_name: id} map."""
    loc_map = {}
    for lib_key, config in LOCATION_MAP.items():
        loc_name = config["name"]
        loc_type = config["type"]

        cur.execute("SELECT id FROM locations WHERE name = %s", (loc_name,))
        row = cur.fetchone()
        if row:
            loc_map[lib_key] = row[0]
        else:
            cur.execute(
                "INSERT INTO locations (name, type) VALUES (%s, %s) RETURNING id",
                (loc_name, loc_type),
            )
            loc_map[lib_key] = cur.fetchone()[0]

    return loc_map


def seed_books(
    cur,
    wb,
    *,
    dry_run: bool = False,
    priorities: set[int] | None = None,
    limit: int | None = None,
) -> dict:
    """Ingest books, filtered by priority. Returns stats dict."""
    ws = wb["Books"] if "Books" in wb.sheetnames else None
    if not ws:
        console.print("[yellow]Sheet 'Books' not found[/yellow]")
        return {}

    if priorities is None:
        priorities = {5, 4, 3}

    stats = {
        "works": 0,
        "editions": 0,
        "instances": 0,
        "skipped_dup": 0,
        "skipped_no_priority": 0,
        "skipped_priority_filter": 0,
        "authors_created": 0,
        "series_created": 0,
        "work_subjects_linked": 0,
        "work_keywords_linked": 0,
    }

    if dry_run:
        # Count qualifying rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            is_dup = clean(_col(row, "duplicated_entry"))
            priority = clean_int(_col(row, "priority"))
            if is_dup == "Y":
                stats["skipped_dup"] += 1
                continue
            if priority is None:
                stats["skipped_no_priority"] += 1
                continue
            if priority not in priorities:
                stats["skipped_priority_filter"] += 1
                continue
            stats["works"] += 1
            if limit and stats["works"] >= limit:
                break
        return stats

    # Pre-load lookups
    author_lookup: dict[str, str] = {}
    cur.execute("SELECT sort_name, id FROM authors WHERE sort_name IS NOT NULL")
    for r in cur.fetchall():
        author_lookup[r[0]] = r[1]
    cur.execute("SELECT name, id FROM authors")
    for r in cur.fetchall():
        author_lookup[r[0]] = r[1]

    series_lookup: dict[str, str] = {}
    cur.execute("SELECT LOWER(title), id FROM series")
    for r in cur.fetchall():
        series_lookup[r[0]] = r[1]

    work_type_lookup = bulk_load_lookup(cur, "work_types")
    subject_lookup = bulk_load_lookup(cur, "subjects")
    keyword_lookup = bulk_load_lookup(cur, "keywords")
    language_lookup: dict[str, str] = {}
    cur.execute("SELECT LOWER(name), iso_639_1 FROM languages WHERE iso_639_1 IS NOT NULL")
    for r in cur.fetchall():
        language_lookup[r[0]] = r[1]
    # Also index by ISO code itself
    cur.execute("SELECT iso_639_1, iso_639_1 FROM languages WHERE iso_639_1 IS NOT NULL")
    for r in cur.fetchall():
        language_lookup[r[0].lower()] = r[1]

    loc_map = _ensure_locations(cur)

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        is_dup = clean(_col(row, "duplicated_entry"))
        if is_dup == "Y":
            stats["skipped_dup"] += 1
            continue

        priority = clean_int(_col(row, "priority"))
        if priority is None:
            stats["skipped_no_priority"] += 1
            continue
        if priority not in priorities:
            stats["skipped_priority_filter"] += 1
            continue

        title = clean(_col(row, "title"))
        if not title:
            continue

        author_raw = clean(_col(row, "author"))
        series_raw = clean(_col(row, "series"))
        series_number = clean(_col(row, "series_number"))
        pub_date = clean(_col(row, "published_date"))
        edition_name = clean(_col(row, "edition"))
        publisher = clean(_col(row, "publisher"))
        pub_title = clean(_col(row, "published_title"))
        pub_language = clean(_col(row, "published_language"))
        page_count = clean_int(_col(row, "page_count"))
        description = clean(_col(row, "description_genai"))
        rating_raw = clean_int(_col(row, "rating"))
        book_type_raw = clean(_col(row, "book_type"))
        temp_tags = clean(_col(row, "temporary_tags"))
        tradition = clean(_col(row, "tradition"))
        notes = clean(_col(row, "notes"))
        cover_raw = clean(_col(row, "cover_raw"))
        goodreads_url = clean(_col(row, "goodreads_url")) or clean(_col(row, "goodreads_path"))
        hardcover_url = clean(_col(row, "hardcover_url"))

        # --- WORK ---
        rating = rating_raw if rating_raw is not None else priority
        catalogue_status = _derive_catalogue_status(row, priority)
        acq_priority = _derive_acquisition_priority(priority)

        # Work type resolution
        work_type_id = None
        if book_type_raw:
            work_type_id = lookup_id_ilike(cur, "work_types", "name", book_type_raw)

        # Series resolution
        series_id = None
        if series_raw:
            initial_count = len(series_lookup)
            series_id = _resolve_or_create_series(cur, series_raw, series_lookup)
            if len(series_lookup) > initial_count:
                stats["series_created"] += 1

        # Language resolution
        lang_code = "en"  # default
        if pub_language:
            lang_code = language_lookup.get(pub_language.lower(), pub_language.lower()[:2] if len(pub_language) > 1 else "en")

        pub_year = _parse_publication_year(pub_date)

        # Use first author for slug generation
        first_author = split_authors(author_raw)[0] if author_raw else None
        work_slug = slugify(f"{title}-by-{invert_author_name(first_author) if first_author else 'unknown'}")
        if not work_slug:
            work_slug = slugify(title)

        # Check for existing work by slug
        existing_work_id = lookup_id(cur, "works", "slug", work_slug)
        if existing_work_id:
            work_id = existing_work_id
        else:
            work_id = upsert_returning_id(
                cur, "works",
                [
                    "title", "slug", "original_language", "original_year",
                    "description", "series_name", "series_position", "series_id",
                    "work_type_id", "notes", "rating",
                    "catalogue_status", "acquisition_priority",
                    "metadata_source",
                ],
                (
                    title, work_slug, lang_code, pub_year,
                    description, series_raw, series_number, series_id,
                    work_type_id, notes, rating,
                    catalogue_status, acq_priority,
                    METADATA_SOURCE,
                ),
                "slug",
            )
        stats["works"] += 1

        # --- AUTHOR LINKING (supports multiple authors separated by &) ---
        if author_raw and work_id:
            author_parts = split_authors(author_raw)
            for sort_order, author_part in enumerate(author_parts):
                initial_count = len(author_lookup)
                author_id = _resolve_or_create_author(cur, author_part, author_lookup)
                if len(author_lookup) > initial_count:
                    stats["authors_created"] += 1
                if author_id:
                    cur.execute(
                        """INSERT INTO work_authors (work_id, author_id, role, sort_order)
                        VALUES (%s, %s, 'author', %s) ON CONFLICT DO NOTHING""",
                        (work_id, author_id, sort_order),
                    )

        # --- EDITION ---
        edition_title = pub_title if pub_title else title
        goodreads_id = extract_goodreads_id(goodreads_url)

        # Determine if translated
        is_translated = False
        if pub_language and lang_code != "en":
            # If published in a different language than the work's original, it's a translation
            # For now, we just mark non-English editions as potentially translated
            pass

        edition_id = None
        if work_id:
            cur.execute(
                """INSERT INTO editions (
                    work_id, title, publisher, publication_year, language,
                    page_count, edition_name, is_translated,
                    cover_source_url, goodreads_id,
                    metadata_source
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id""",
                (
                    work_id, edition_title, publisher, pub_year, lang_code,
                    page_count, edition_name, is_translated,
                    cover_raw, goodreads_id,
                    METADATA_SOURCE,
                ),
            )
            edition_id = cur.fetchone()[0]
            stats["editions"] += 1

        # --- INSTANCES ---
        if edition_id:
            for lib_key, col_idx in LIB_COLUMNS.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is None:
                    continue
                val_str = str(val).strip()
                if not val_str or val_str.upper() == "N":
                    continue

                loc_id = loc_map.get(lib_key)
                if not loc_id:
                    continue

                fmt = LOCATION_MAP[lib_key]["format"]
                cur.execute(
                    """INSERT INTO instances (edition_id, location_id, format, status)
                    VALUES (%s, %s, %s, 'available')""",
                    (edition_id, loc_id, fmt),
                )
                stats["instances"] += 1

        # --- TAXONOMY LINKING ---
        if work_id and tradition:
            traditions = parse_semicolons(tradition)
            for t in traditions:
                subj_id = subject_lookup.get(t)
                if not subj_id:
                    subj_id = lookup_id_ilike(cur, "subjects", "name", t)
                if subj_id:
                    insert_junction(cur, "work_subjects", "work_id", "subject_id", work_id, subj_id)
                    stats["work_subjects_linked"] += 1

        if work_id and temp_tags:
            tags = parse_semicolons(temp_tags)
            for t in tags:
                kw_id = keyword_lookup.get(t)
                if not kw_id:
                    kw_id = lookup_id_ilike(cur, "keywords", "name", t)
                if kw_id:
                    insert_junction(cur, "work_keywords", "work_id", "keyword_id", work_id, kw_id)
                    stats["work_keywords_linked"] += 1

        count += 1
        if limit and count >= limit:
            break

        # Progress every 100 rows
        if count % 100 == 0:
            console.print(f"  ... processed {count} books")

    return stats


def run(*, dry_run: bool = False, priorities: set[int] | None = None, limit: int | None = None):
    console.rule("[bold]Step 5: Books Inventory")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    if priorities is None:
        priorities = {5, 4, 3}

    console.print(f"  Filtering to priorities: {sorted(priorities, reverse=True)}")

    try:
        if dry_run:
            console.print("[yellow]DRY RUN — no data will be written[/yellow]\n")
            stats = seed_books(None, wb, dry_run=True, priorities=priorities, limit=limit)
        else:
            with transaction() as cur:
                stats = seed_books(cur, wb, priorities=priorities, limit=limit)

        console.print(f"\n  works created:           {stats.get('works', 0)}")
        console.print(f"  editions created:        {stats.get('editions', 0)}")
        console.print(f"  instances created:       {stats.get('instances', 0)}")
        console.print(f"  authors auto-created:    {stats.get('authors_created', 0)}")
        console.print(f"  series auto-created:     {stats.get('series_created', 0)}")
        console.print(f"  work-subject links:      {stats.get('work_subjects_linked', 0)}")
        console.print(f"  work-keyword links:      {stats.get('work_keywords_linked', 0)}")
        console.print(f"  skipped (duplicate):     {stats.get('skipped_dup', 0)}")
        console.print(f"  skipped (no priority):   {stats.get('skipped_no_priority', 0)}")
        console.print(f"  skipped (filtered out):  {stats.get('skipped_priority_filter', 0)}")
        console.print("[green]Books inventory complete.[/green]\n")

    finally:
        wb.close()


if __name__ == "__main__":
    import sys
    prio = {5, 4, 3}
    if "--priority" in sys.argv:
        idx = sys.argv.index("--priority") + 1
        prio = {int(x) for x in sys.argv[idx].split(",")}
    run(
        dry_run="--dry-run" in sys.argv,
        priorities=prio,
        limit=int(sys.argv[sys.argv.index("--limit") + 1]) if "--limit" in sys.argv else None,
    )
